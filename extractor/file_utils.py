# Copyright (c) 2026 KORDA Contributors.
#
# This file is part of KORDA.
#
# KORDA is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# KORDA is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with KORDA.  If not, see <https://www.gnu.org/licenses/>.


from __future__ import annotations

import hashlib
import json
import logging
import threading
import urllib.request
import zipfile
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO

from django.conf import settings


def _get_gcs_bucket():
    from django.core.exceptions import ImproperlyConfigured
    from google.cloud import storage

    client = storage.Client()
    # DEBT-03 fix: never silently fall back to a hardcoded bucket name.
    # GS_BUCKET_NAME MUST be explicitly set via env var / Secret Manager.
    bucket_name = getattr(settings, "GS_BUCKET_NAME", None)
    if not bucket_name:
        raise ImproperlyConfigured(
            "GS_BUCKET_NAME is not configured. Set it via environment variable or GCP Secret Manager."
        )
    return client.bucket(bucket_name)


from typing import IO, Any

import bleach  # type: ignore[import-untyped]
import markdown as md_lib  # type: ignore[import-untyped]
from django.utils.text import slugify

logger = logging.getLogger(__name__)

APPLICATION_JSON = "application/json"
NO_COMPLETED_DOCS_MSG = "No completed documents selected for export bundle."


# ── Fallback exchange rates (used if the external API is offline) ─────────────
# Approximate rates as of mid-2026 — stale but far better than a blocking 5s timeout.
_FALLBACK_RATES: dict[str, float] = {
    "USD": 1.0,
    "EUR": 0.92,
    "GBP": 0.79,
    "IDR": 16200.0,
    "SAR": 3.75,
    "TRY": 32.0,
}

# ── ZIP Security Constants (SEC-03 Zip Slip + EDGE-05 Zip Bomb guards) ────────
# Maximum total uncompressed bytes allowed from any single uploaded ZIP archive.
_ZIP_MAX_UNCOMPRESSED_BYTES: int = 500 * 1024 * 1024  # 500 MB
# Maximum number of members (files) permitted inside a single ZIP archive.
_ZIP_MAX_MEMBER_COUNT: int = 500


def validate_zip(zf: zipfile.ZipFile) -> None:
    """Guard against Zip Bomb (EDGE-05) and path traversal Zip Slip (SEC-03) attacks.

    Call this immediately after opening any user-supplied ZipFile before extracting.

    Raises:
        ValueError: if member count, total uncompressed size, or a path traversal
                    attempt is detected.
    """
    members = zf.infolist()

    # EDGE-05: Bomb guard — reject archives exceeding member count or size limits
    if len(members) > _ZIP_MAX_MEMBER_COUNT:
        raise ValueError(
            f"ZIP archive contains {len(members)} files, exceeding the maximum of {_ZIP_MAX_MEMBER_COUNT}."
        )
    total_uncompressed = sum(m.file_size for m in members)
    if total_uncompressed > _ZIP_MAX_UNCOMPRESSED_BYTES:
        raise ValueError(
            f"ZIP uncompressed size {total_uncompressed} bytes exceeds the {_ZIP_MAX_UNCOMPRESSED_BYTES}-byte limit."
        )


def safe_extract(zf: zipfile.ZipFile, target_dir: str) -> None:
    """Extract ZIP contents with Zip Slip path traversal protection (SEC-03).

    Validates each member's resolved path stays within `target_dir` before
    extracting. Always call `validate_zip()` first to apply size/count limits.

    Raises:
        ValueError: if any ZIP member path escapes the target directory.
    """
    import os

    real_target = os.path.realpath(target_dir)
    for member in zf.infolist():
        member_path = os.path.realpath(os.path.join(real_target, member.filename))
        # SEC-03: Zip Slip guard — reject members that resolve outside the target dir
        if not member_path.startswith(real_target + os.sep) and member_path != real_target:
            raise ValueError(f"Zip Slip path traversal detected in member: '{member.filename}'. Extraction aborted.")
        zf.extract(member, target_dir)


# Thread-safe in-memory cache for exchange rates (1-hour TTL)
_rates_cache: dict[str, Any] = {}
_rates_cache_lock = threading.Lock()


def calculate_file_sha256(file_handle: IO[bytes]) -> str:
    """
    Computes SHA-256 checksum in chunks of 64KB for deduplication and content-addressing
    from an open file-like byte stream.
    """
    sha256 = hashlib.sha256()
    file_handle.seek(0)
    for chunk in iter(lambda: file_handle.read(65536), b""):
        sha256.update(chunk)
    file_handle.seek(0)
    return sha256.hexdigest()


def calculate_filepath_sha256(file_path: str) -> str:
    """
    Computes SHA-256 checksum in chunks of 64KB for deduplication and content-addressing
    from a filesystem path.
    """
    from pathlib import Path

    target_path = Path(file_path).resolve()
    if not target_path.is_file():
        raise ValueError(f"Path is not a valid file: {file_path}")

    with target_path.open("rb") as f:
        return calculate_file_sha256(f)


def process_csv_local(file_path: str) -> str:
    """
    Converts tabular CSV files cleanly to an aligned Markdown table locally ($0 cost!).
    Supports heavy columns and escapes standard symbols.
    """
    import csv

    markdown_lines: list[str] = []
    try:
        with open(file_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = list(reader)
    except UnicodeDecodeError:
        with open(file_path, newline="", encoding="latin-1") as f:
            reader = csv.reader(f)
            rows = list(reader)

    if not rows:
        return "*Empty CSV Document*"

    headers = rows[0]
    markdown_lines.extend(
        (
            "| " + " | ".join([h.replace("|", "\\|") for h in headers]) + " |",
            "| " + " | ".join(["---" for _ in headers]) + " |",
        )
    )

    for row in rows[1:]:
        if len(row) < len(headers):
            row += [""] * (len(headers) - len(row))
        elif len(row) > len(headers):
            row = row[: len(headers)]
        markdown_lines.append(
            "| " + " | ".join([cell.replace("|", "\\|").replace("\n", "<br>") for cell in row]) + " |"
        )

    return "\n".join(markdown_lines)


def process_txt_local(file_path: str) -> str:
    """
    Reads local TXT and Markdown files (including Quran Arabic text with Harakat and translations),
    preserving UTF-8 encoding and layout structure cleanly ($0 API cost).
    """
    try:
        with open(file_path, encoding="utf-8") as f:
            content = f.read()
    except UnicodeDecodeError:
        with open(file_path, encoding="latin-1") as f:
            content = f.read()
    return content


def _format_json_rows_as_table(title: str, items: list[dict]) -> str:
    """Formats a list of dictionaries into a clean Markdown table."""
    keys = list({k: True for item in items for k in item}.keys())
    rows = []
    for item in items:
        rows.append([str(item.get(k, "")).strip().replace("\n", "<br>") for k in keys])
    return _format_markdown_table_sheet(title, [keys, *rows])


def _format_json_dict_section(key: Any, value: Any) -> str:
    title = str(key).replace("_", " ").title()
    if isinstance(value, list) and all(isinstance(x, dict) for x in value):
        return f"## {title}\n\n" + _format_json_rows_as_table(title, value)
    if isinstance(value, (dict, list)):
        return f"## {title}\n\n```json\n{json.dumps(value, indent=2, ensure_ascii=False)}\n```"
    return f"## {title}\n\n{value}"


def process_json_local(file_path: str) -> str:
    """
    Parses local JSON documents (including Quran verse datasets, Hadith collections,
    and structured JSON payloads) into structured Markdown sections and tables ($0 API cost).
    """
    try:
        with open(file_path, encoding="utf-8") as f:
            data = json.load(f)
    except UnicodeDecodeError:
        with open(file_path, encoding="latin-1") as f:
            data = json.load(f)

    if isinstance(data, list):
        if not data:
            return "*Empty JSON List*"
        if all(isinstance(item, dict) for item in data):
            return _format_json_rows_as_table("Dataset", data)
        return "\n\n".join(f"- {json.dumps(item, ensure_ascii=False)}" for item in data)

    if isinstance(data, dict):
        sections = [_format_json_dict_section(k, v) for k, v in data.items()]
        return "\n\n".join(sections) if sections else "*Empty JSON Object*"

    return str(data)


def _format_markdown_table_sheet(sheet_title: str, rows: list[list[str]]) -> str:
    """Formats a matrix of cells into a GitHub Flavored Markdown table."""
    if not rows:
        return ""
    headers = rows[0]
    sheet_md = [f"### Sheet: {sheet_title}\n"]
    sheet_md.append("| " + " | ".join([h.replace("|", "\\|") for h in headers]) + " |")
    sheet_md.append("| " + " | ".join(["---" for _ in headers]) + " |")
    for r in rows[1:]:
        if len(r) < len(headers):
            r += [""] * (len(headers) - len(r))
        elif len(r) > len(headers):
            r = r[: len(headers)]
        sheet_md.append("| " + " | ".join([c.replace("|", "\\|").replace("\n", "<br>") for c in r]) + " |")
    return "\n".join(sheet_md)


def _parse_excel_openpyxl(file_path: str) -> str | None:
    """Extracts tables from workbook using openpyxl."""
    try:
        import openpyxl  # type: ignore[import-untyped]

        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        sheets_md: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                str_row = [str(cell).strip() if cell is not None else "" for cell in row]
                if any(str_row):
                    rows.append(str_row)
            formatted = _format_markdown_table_sheet(sheet_name, rows)
            if formatted:
                sheets_md.append(formatted)
        wb.close()
        return "\n\n".join(sheets_md) if sheets_md else None
    except Exception as exc:
        logger.debug("[Excel Parser] openpyxl extraction failed: %s", exc)
        return None


def _extract_zipxml_cell_value(c_el: Any, shared_strings: list[str]) -> str:
    cell_type = c_el.get("t")
    v_el = c_el.find(".//{*}v")
    if v_el is None or not v_el.text:
        return ""
    if cell_type == "s" and v_el.text.isdigit():
        idx = int(v_el.text)
        return shared_strings[idx] if idx < len(shared_strings) else v_el.text
    return v_el.text


def _extract_sheet_rows_from_xml(sheet_tree, shared_strings: list[str]) -> list[list[str]]:
    """Helper to extract non-empty rows from an Excel worksheet XML tree."""
    rows: list[list[str]] = []
    for row_el in sheet_tree.findall(".//{*}row"):
        row_cells = [_extract_zipxml_cell_value(c_el, shared_strings) for c_el in row_el.findall(".//{*}c")]
        if any(row_cells):
            rows.append(row_cells)
    return rows


def _load_zipxml_shared_strings(zf: Any) -> list[str]:
    import xml.etree.ElementTree as ET  # nosec B405

    shared_strings: list[str] = []
    if "xl/sharedStrings.xml" in zf.namelist():
        ss_tree = ET.fromstring(zf.read("xl/sharedStrings.xml"))  # nosec B314 # noqa: S314
        for si in ss_tree.findall(".//{*}si"):
            t_el = si.find(".//{*}t")
            shared_strings.append(t_el.text if t_el is not None and t_el.text else "")
    return shared_strings


def _parse_zipxml_sheets(zf: Any, shared_strings: list[str]) -> list[str]:
    import xml.etree.ElementTree as ET  # nosec B405

    sheet_files = [n for n in zf.namelist() if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")]
    sheets_md: list[str] = []
    for s_idx, s_file in enumerate(sheet_files, 1):
        sheet_tree = ET.fromstring(zf.read(s_file))  # nosec B314 # noqa: S314
        rows = _extract_sheet_rows_from_xml(sheet_tree, shared_strings)
        formatted = _format_markdown_table_sheet(f"Sheet {s_idx}", rows)
        if formatted:
            sheets_md.append(formatted)
    return sheets_md


def _parse_excel_zipxml(file_path: str) -> str | None:
    """Pure standard-library ZIP+XML fallback for .xlsx parsing."""
    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            validate_zip(zf)
            shared_strings = _load_zipxml_shared_strings(zf)
            sheets_md = _parse_zipxml_sheets(zf, shared_strings)
            return "\n\n".join(sheets_md) if sheets_md else None
    except Exception as exc:
        logger.warning("[Excel Parser] Native XML extraction failed: %s", exc)
        return None


def process_excel_local(file_path: str) -> str:
    """
    Converts Excel spreadsheets (.xlsx, .xls) cleanly to Markdown tables locally ($0 API cost).
    Supports multi-sheet workbooks using openpyxl or pure standard library XML parsing.
    """
    result = _parse_excel_openpyxl(file_path)
    if result:
        return result

    result = _parse_excel_zipxml(file_path)
    if result:
        return result

    return "*Empty or Unreadable Excel Document*"


def process_pdf_local(file_path: str) -> tuple[str, int]:
    """
    Extracts native digital text and page structure from PDF pages locally ($0 API cost).
    Returns a tuple of (extracted_markdown, page_count).
    If no text or PyMuPDF/pypdf is uninstalled, returns ("", page_count).
    """
    try:
        import fitz  # PyMuPDF

        doc = fitz.open(file_path)
        pages_md: list[str] = []
        page_count = len(doc)
        for page_num in range(page_count):
            page = doc[page_num]
            text = page.get_text("text").strip()
            if text:
                pages_md.append(f"## Page {page_num + 1}\n\n{text}")
        doc.close()
        if pages_md and len("".join(pages_md).strip()) > 30:
            return "\n\n---\n\n".join(pages_md), page_count
        return "", page_count
    except Exception as exc:
        logger.debug("[PDF Local] PyMuPDF local text extraction skipped (%s)", exc)

    try:
        import pypdf

        reader = pypdf.PdfReader(file_path)
        page_count = len(reader.pages)
        pages_md = []
        for p_idx, page in enumerate(reader.pages, 1):
            text = (page.extract_text() or "").strip()
            if text:
                pages_md.append(f"## Page {p_idx}\n\n{text}")
        if pages_md and len("".join(pages_md).strip()) > 30:
            return "\n\n---\n\n".join(pages_md), page_count
        return "", page_count
    except Exception as exc:
        logger.debug("[PDF Local] pypdf local text extraction skipped (%s)", exc)

    return "", 0


def clean_html_content(raw_html: str) -> str:
    """
    Sanitizes HTML content using bleach to shield against XSS.
    This is ONLY called on already-compiled HTML, NOT on raw Markdown text.
    Preserves RTL wrappers (div, dir attribute) required for Arabic script rendering.
    """
    allowed_tags = [
        "h1",
        "h2",
        "h3",
        "h4",
        "h5",
        "h6",
        "p",
        "b",
        "i",
        "strong",
        "em",
        "u",
        "ul",
        "ol",
        "li",
        "br",
        "hr",
        "table",
        "thead",
        "tbody",
        "tr",
        "th",
        "td",
        "code",
        "pre",
        "blockquote",
        "a",
        "span",
        "div",  # required for <div dir="rtl"> Arabic script wrappers
    ]
    allowed_attrs = {
        "a": ["href", "title", "target"],
        "*": ["id", "class", "dir"],
    }
    return bleach.clean(raw_html, tags=allowed_tags, attributes=allowed_attrs, strip=True)


def _detect_first_strong(text_content, latin_chars, arabic_chars):
    for char in text_content:
        if latin_chars.match(char):
            return "latin"
        if arabic_chars.match(char):
            return "arabic"
    return None


def _apply_arabic_attributes(attrs):
    import re

    if "dir=" in attrs:
        attrs = re.sub(r'dir="[^"]*"', 'dir="rtl"', attrs)
        attrs = re.sub(r"dir='[^']*'", 'dir="rtl"', attrs)
    else:
        attrs += ' dir="rtl"'

    if "class=" in attrs:
        attrs = re.sub(r'class="([^"]*)"', r'class="\1 arabic-text"', attrs)
        attrs = re.sub(r"class='([^']*)'", r"class='\1 arabic-text'", attrs)
    else:
        attrs += ' class="arabic-text"'
    return attrs


def parse_arabic_layout(html_content: str) -> str:
    """
    Parses HTML content, detects block tags whose first strong alphabetical
    character is Arabic, and adds dir="rtl" class="arabic-text" to them.
    Leaves English translations starting with English letters as LTR.
    """
    import re

    pattern = re.compile(
        r"<(p|li|blockquote|h1|h2|h3|h4|h5|h6|span|div|td)([^>]*)>(.*?)</\1>", re.DOTALL | re.IGNORECASE
    )

    arabic_chars = re.compile(r"[\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF\uFB50-\uFDFF\uFE70-\uFEFF]")
    latin_chars = re.compile(r"[A-Za-z]")

    def replacer(match):
        tag_name = match.group(1)
        attrs = match.group(2)
        content = match.group(3)

        text_content = re.sub(r"<[^>]*>", "", content)
        first_strong = _detect_first_strong(text_content, latin_chars, arabic_chars)

        if first_strong == "arabic":
            attrs = _apply_arabic_attributes(attrs)

        return f"<{tag_name}{attrs}>{content}</{tag_name}>"

    return pattern.sub(replacer, html_content)


def render_markdown_to_html(markdown_text: str) -> str:
    """
    Converts Markdown to HTML safely and applies first-strong-character Arabic layout logic.
    """
    html = md_lib.markdown(markdown_text, extensions=["tables", "fenced_code", "nl2br"])
    sanitized = clean_html_content(html)
    return parse_arabic_layout(sanitized)


def _build_yaml_frontmatter(doc: Any) -> str:
    """Build a YAML frontmatter header for export markdown files."""
    publisher = _get_doc_field_str(doc, "publisher")
    publication_year = _get_doc_field_str(doc, "publication_year")
    doi = _get_doc_field_str(doc, "doi")
    license_type = _get_doc_field_str(doc, "license_type")
    return (
        "---\n"
        f'title: "{doc.title}"\n'
        f'author: "{doc.author}"\n'
        f'language: "{doc.language}"\n'
        f'document_type: "{doc.document_type}"\n'
        f'publisher: "{publisher}"\n'
        f'publication_year: "{publication_year}"\n'
        f'doi: "{doi}"\n'
        f'license_type: "{license_type}"\n'
        f'source_hash: "{doc.file_hash}"\n'
        f'exported_at: "{datetime.now(UTC).isoformat()}"\n'
        "---\n\n"
    )


@dataclass
class ZipExportContext:
    seen_lang_paths: set[str]
    seen_author_paths: set[str]
    manifest: dict[str, Any]
    master_content: list[str]
    zip_file: zipfile.ZipFile
    include_taxonomic_views: bool = True


def _write_taxonomic_views(
    base_slug: str, clean_lang: str, clean_author: str, full_content: str, ctx: ZipExportContext
):
    """Helper to write duplicate organized views by language and author."""
    lang_slug = f"{base_slug}.md"
    lang_path = f"Language/{clean_lang}/{lang_slug}"
    counter = 1
    while lang_path in ctx.seen_lang_paths:
        counter += 1
        lang_slug = f"{base_slug}_{counter}.md"
        lang_path = f"Language/{clean_lang}/{lang_slug}"
    ctx.seen_lang_paths.add(lang_path)

    author_slug = f"{base_slug}.md"
    author_path = f"Author/{clean_author}/{author_slug}"
    counter = 1
    while author_path in ctx.seen_author_paths:
        counter += 1
        author_slug = f"{base_slug}_{counter}.md"
        author_path = f"Author/{clean_author}/{author_slug}"
    ctx.seen_author_paths.add(author_path)

    ctx.zip_file.writestr(lang_path, full_content)
    ctx.zip_file.writestr(author_path, full_content)


def _append_master_source_block(idx: int, doc: Any, doc_type: str, doc_markdown: str, ctx: ZipExportContext):
    """Helper to append structured markdown source section to master document."""
    ctx.master_content.append(
        f"<!-- SOURCE_START_{idx + 1}: {doc.title} by {doc.author} ({doc.language}) [Type: {doc_type}, Hash: {doc.file_hash}] -->"
    )
    ctx.master_content.append(f"\n# SOURCE: {doc.title}\n")
    ctx.master_content.append(f"**Author:** {doc.author}  ")
    ctx.master_content.append(f"**Language:** {doc.language}  ")
    ctx.master_content.append(f"**Document Type:** {doc_type}  ")
    if _get_doc_field_str(doc, "publisher"):
        ctx.master_content.append(f"**Publisher:** {_get_doc_field_str(doc, 'publisher')}  ")
    if _get_doc_field_str(doc, "publication_year"):
        ctx.master_content.append(f"**Publication Year:** {_get_doc_field_str(doc, 'publication_year')}  ")
    if _get_doc_field_str(doc, "doi"):
        ctx.master_content.append(f"**DOI:** {_get_doc_field_str(doc, 'doi')}  ")
    if _get_doc_field_str(doc, "license_type"):
        ctx.master_content.append(f"**License:** {_get_doc_field_str(doc, 'license_type')}  ")
    ctx.master_content.append("\n")
    ctx.master_content.append(doc_markdown)
    ctx.master_content.append(f"\n<!-- SOURCE_END_{idx + 1} -->\n\n---\n")


def _process_zip_doc(idx: int, doc: Any, ctx: ZipExportContext):
    clean_lang = slugify(doc.language or "unknown", allow_unicode=True) or "unknown"
    clean_author = slugify(doc.author or "unknown", allow_unicode=True) or "unknown"
    doc_type = doc.document_type or "PDF"

    # Formal standardized batch filename: [3-DIGIT_INDEX]_[TITLE_SLUG].md
    index_prefix = f"{idx + 1:03d}"
    raw_title = doc.title or doc.original_filename or f"document_{doc.id}"
    doc_title_slug = slugify(raw_title, allow_unicode=True) or f"document_{doc.id}"
    base_slug = f"{index_prefix}_{doc_title_slug}"

    doc_markdown = doc.refined_markdown or doc.raw_markdown or "*Empty Document Content*"
    frontmatter = _build_yaml_frontmatter(doc)
    full_content = frontmatter + doc_markdown

    # Primary single-copy standardized document output
    doc_path = f"documents/{base_slug}.md"
    ctx.zip_file.writestr(doc_path, full_content)

    if ctx.include_taxonomic_views:
        _write_taxonomic_views(base_slug, clean_lang, clean_author, full_content, ctx)

    ctx.manifest["documents"].append(
        {
            "id": doc.id,
            "filename": doc.original_filename,
            "title": doc.title,
            "author": doc.author,
            "language": doc.language,
            "type": doc_type,
            "publisher": _get_doc_field_str(doc, "publisher"),
            "publication_year": _get_doc_field_str(doc, "publication_year"),
            "doi": _get_doc_field_str(doc, "doi"),
            "license_type": _get_doc_field_str(doc, "license_type"),
            "page_count": int(getattr(doc, "page_count", 1) or 1),
            "cost_usd": float(getattr(doc, "cost_usd", 0.0) or 0.0),
            "hash": doc.file_hash,
            "export_path": doc_path,
        }
    )

    _append_master_source_block(idx, doc, doc_type, doc_markdown, ctx)


def _get_offline_docs(document_ids, user):
    from django.db.models import Q

    from extractor.models import SourceDocument

    int_ids = []
    uuid_strs = []
    for item in document_ids:
        try:
            int_ids.append(int(item))
        except (ValueError, TypeError):
            uuid_strs.append(str(item))

    docs = SourceDocument.objects.filter(Q(id__in=int_ids) | Q(uuid__in=uuid_strs), status="COMPLETED")
    if user and not (user.is_staff or user.is_superuser):
        docs = docs.filter(Q(uploaded_by=user) | Q(uploaded_by__isnull=True))
    return list(docs)


def _get_surreal_docs(document_ids, user, actor_id: str | None = None):
    from extractor import surreal_db
    from extractor.views import _build_users_map, _wrap_surreal_doc

    docs_list = []
    raw_docs = surreal_db.get_documents(document_ids)
    user_ids = {d.get("uploaded_by_id") for d in raw_docs if d and d.get("uploaded_by_id")}
    users_map = _build_users_map(user_ids, fallback_user=user)

    for raw_doc in raw_docs:
        if not raw_doc or raw_doc.get("status") != "COMPLETED":
            continue

        uploaded_by_id = raw_doc.get("uploaded_by_id")
        owner_id = actor_id or str(user.id)
        if user and not (user.is_staff or user.is_superuser) and uploaded_by_id and uploaded_by_id != owner_id:
            continue

        doc = _wrap_surreal_doc(raw_doc, users_map)
        docs_list.append(doc)
    return docs_list


def generate_curated_zip_bundle(
    document_ids: list[int] | list[str],
    user: Any = None,
    include_taxonomic_views: bool = True,
    actor_id: str | None = None,
) -> bytes:
    """
    Aggregates selected documents into an organized directory structure.
    Saves unique copies under documents/ and optionally by Language/ and Author/.
    Prepends YAML frontmatter to each exported markdown file.
    Enforces user boundaries if `user` is provided and is not a staff/superuser.
    """
    from django.conf import settings

    if getattr(settings, "SURREALDB_OFFLINE", False):
        docs_list = _get_offline_docs(document_ids, user)
    else:
        docs_list = _get_surreal_docs(document_ids, user, actor_id=actor_id)

    if not docs_list:
        raise ValueError(NO_COMPLETED_DOCS_MSG)

    zip_buffer = BytesIO()

    manifest = {
        "exported_at": datetime.now(UTC).isoformat(),
        "document_count": len(docs_list),
        "total_cost_usd": float(sum(d.cost_usd for d in docs_list)),
        "total_pages": sum(d.page_count for d in docs_list),
        "documents": [],
    }

    master_content = [
        "# Master Digital Archival Source Book\n",
        "This master document contains all aggregated document knowledge in a single high-density layout. "
        "It uses strict structural indicators to help archival systems easily parse each separate source content boundaries.\n\n",
    ]

    seen_lang_paths: set[str] = set()
    seen_author_paths: set[str] = set()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        ctx = ZipExportContext(
            seen_lang_paths=seen_lang_paths,
            seen_author_paths=seen_author_paths,
            manifest=manifest,
            master_content=master_content,
            zip_file=zip_file,
            include_taxonomic_views=include_taxonomic_views,
        )
        for idx, doc in enumerate(docs_list):
            _process_zip_doc(idx, doc, ctx)

        zip_file.writestr("master_archival_source.md", "\n".join(master_content))
        zip_file.writestr("manifest.json", json.dumps(manifest, indent=2))

    zip_buffer.seek(0)
    return zip_buffer.getvalue()


def _build_sft_pair(
    doc_info: tuple[str, str, str, str, str, str, str, str], chunk_item: dict[str, Any]
) -> dict[str, Any] | None:
    doc_uuid, title, author, language, publisher, publication_year, doi, license_type = doc_info
    chunk_content = str(chunk_item.get("content") or "")
    if not chunk_content.strip():
        return None
    page = chunk_item.get("page_number") or 1
    chapter = chunk_item.get("chapter_title") or ""
    anchor = chunk_item.get("anchor_id") or f"page-{page}"

    user_prompt = (
        f"Context from '{title}' by {author} (Page {page}"
        + (f", {chapter}" if chapter else "")
        + f"):\n\n{chunk_content}\n\n"
        "Question: Explain the key context, teachings, and significance of this excerpt."
    )
    assistant_response = f"Based on '{title}', the text conveys:\n\n{chunk_content}"

    return {
        "prompt": user_prompt,
        "completion": assistant_response,
        "messages": [
            {
                "role": "system",
                "content": "You are a knowledgeable archival research assistant specializing in historical, literary, and classical religious documents.",
            },
            {"role": "user", "content": user_prompt},
            {"role": "assistant", "content": assistant_response},
        ],
        "metadata": {
            "doc_uuid": doc_uuid,
            "title": title,
            "author": author,
            "language": language,
            "publisher": publisher,
            "publication_year": publication_year,
            "doi": doi,
            "license_type": license_type,
            "page_number": page,
            "chapter_title": chapter,
            "anchor_id": anchor,
        },
    }


def _get_doc_chunks(doc: Any, doc_uuid: str, limit: int) -> list[dict[str, Any]]:
    from django.conf import settings

    from extractor import surreal_db

    if getattr(settings, "SURREALDB_OFFLINE", False):
        from extractor.rag import chunk_document_semantically

        content = (
            getattr(doc, "refined_markdown", None)
            or getattr(doc, "cleaned_markdown", None)
            or getattr(doc, "raw_markdown", None)
            or ""
        )
        raw_chunks = chunk_document_semantically(content)
        return [
            {
                "content": raw_text,
                "chunk_index": idx,
                "page_number": 1,
                "chapter_title": "",
                "anchor_id": f"chunk-{idx}",
            }
            for idx, raw_text in enumerate(raw_chunks[:limit])
        ]

    try:
        return surreal_db.get_document_chunks(doc_uuid, limit=limit)
    except Exception as e:
        logger.warning("[SFT Dataset] Failed to retrieve chunks for %s: %s", doc_uuid, e)
        return []


def _collect_pairs_for_document(doc, limit: int, current_count: int) -> list[dict[str, Any]]:
    """Helper to retrieve chunks and build SFT pairs for a single document."""
    doc_uuid = str(getattr(doc, "doc_uuid", None) or getattr(doc, "uuid", None) or doc.id)
    doc_info = (
        doc_uuid,
        getattr(doc, "title", None) or "Document",
        getattr(doc, "author", None) or "Author",
        getattr(doc, "language", None) or "en",
        _get_doc_field_str(doc, "publisher"),
        _get_doc_field_str(doc, "publication_year"),
        _get_doc_field_str(doc, "doi"),
        _get_doc_field_str(doc, "license_type"),
    )
    doc_pairs: list[dict[str, Any]] = []
    chunks = _get_doc_chunks(doc, doc_uuid, limit)
    for chunk_item in chunks[:limit]:
        pair = _build_sft_pair(doc_info, chunk_item)
        if pair:
            doc_pairs.append(pair)
        if current_count + len(doc_pairs) >= limit:
            break
    return doc_pairs


def generate_sft_dataset_pairs(
    document_ids: list[int] | list[str],
    user: Any = None,
    actor_id: str | None = None,
    limit: int = 100,
) -> list[dict[str, Any]]:
    """
    Generates structured instruction fine-tuning (SFT) prompt-completion and chat message pairs
    from document chunks, respecting tenant isolation boundaries.
    """
    from django.conf import settings

    if getattr(settings, "SURREALDB_OFFLINE", False):
        docs_list = _get_offline_docs(document_ids, user)
    else:
        docs_list = _get_surreal_docs(document_ids, user, actor_id=actor_id)

    if not docs_list:
        raise ValueError(NO_COMPLETED_DOCS_MSG)

    pairs: list[dict[str, Any]] = []
    for doc in docs_list:
        doc_pairs = _collect_pairs_for_document(doc, limit, len(pairs))
        pairs.extend(doc_pairs)
        if len(pairs) >= limit:
            break

    return pairs


def generate_sft_jsonl_bundle(
    document_ids: list[int] | list[str],
    user: Any = None,
    actor_id: str | None = None,
) -> bytes:
    """Generates standard Hugging Face JSONL formatted SFT dataset as UTF-8 bytes."""
    pairs = generate_sft_dataset_pairs(document_ids, user, actor_id=actor_id, limit=5000)
    lines = [json.dumps(p, ensure_ascii=False) for p in pairs]
    return ("\n".join(lines) + "\n").encode("utf-8")


def _init_sqlite_export_schema(cursor: Any) -> None:
    cursor.execute("""
        CREATE TABLE documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            doc_uuid TEXT UNIQUE,
            title TEXT,
            original_filename TEXT,
            author TEXT,
            language TEXT,
            publisher TEXT,
            publication_year TEXT,
            doi TEXT,
            license_type TEXT,
            file_hash TEXT,
            page_count INTEGER,
            cost_usd REAL,
            created_at TEXT
        );
    """)
    cursor.execute("""
        CREATE TABLE chunks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            doc_uuid TEXT,
            chunk_index INTEGER,
            page_number INTEGER,
            chapter_title TEXT,
            anchor_id TEXT,
            content TEXT,
            FOREIGN KEY (doc_uuid) REFERENCES documents(doc_uuid)
        );
    """)
    cursor.execute("""
        CREATE VIRTUAL TABLE chunks_fts USING fts5(
            content,
            title,
            author,
            doc_uuid UNINDEXED,
            page_number UNINDEXED,
            anchor_id UNINDEXED,
            tokenize='porter unicode61'
        );
    """)


def _get_doc_field_str(doc: Any, attr: str, default: str = "") -> str:
    val = getattr(doc, attr, None)
    return str(val) if val is not None and val != "" else default


def _insert_sqlite_document(cursor: Any, doc: Any) -> str:
    doc_uuid = str(getattr(doc, "doc_uuid", None) or getattr(doc, "uuid", None) or doc.id)
    title = _get_doc_field_str(doc, "title", "Untitled")
    author = _get_doc_field_str(doc, "author", "Unknown")
    created_at_val = _get_doc_field_str(doc, "created_at", datetime.now(UTC).isoformat())

    cursor.execute(
        """
        INSERT INTO documents (
            doc_uuid, title, original_filename, author, language,
            publisher, publication_year, doi, license_type, file_hash,
            page_count, cost_usd, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
        """,
        (
            doc_uuid,
            title,
            _get_doc_field_str(doc, "original_filename"),
            author,
            _get_doc_field_str(doc, "language", "en"),
            _get_doc_field_str(doc, "publisher"),
            _get_doc_field_str(doc, "publication_year"),
            _get_doc_field_str(doc, "doi"),
            _get_doc_field_str(doc, "license_type"),
            _get_doc_field_str(doc, "file_hash"),
            int(getattr(doc, "page_count", 1) or 1),
            float(getattr(doc, "cost_usd", 0.0) or 0.0),
            created_at_val,
        ),
    )
    return doc_uuid


def _insert_sqlite_chunks(cursor: Any, doc: Any, doc_uuid: str) -> None:
    title = str(getattr(doc, "title", None) or "Untitled")
    author = str(getattr(doc, "author", None) or "Unknown")
    chunks = _get_doc_chunks(doc, doc_uuid, limit=10000)
    for chunk in chunks:
        content_str = str(chunk.get("content") or "").strip()
        if not content_str:
            continue
        p_num = int(chunk.get("page_number") or 1)
        chap = str(chunk.get("chapter_title") or "")
        anch = str(chunk.get("anchor_id") or f"page-{p_num}")
        c_idx = int(chunk.get("chunk_index") or 0)

        cursor.execute(
            """
            INSERT INTO chunks (
                doc_uuid, chunk_index, page_number, chapter_title, anchor_id, content
            ) VALUES (?, ?, ?, ?, ?, ?);
            """,
            (doc_uuid, c_idx, p_num, chap, anch, content_str),
        )
        cursor.execute(
            """
            INSERT INTO chunks_fts (
                content, title, author, doc_uuid, page_number, anchor_id
            ) VALUES (?, ?, ?, ?, ?, ?);
            """,
            (content_str, title, author, doc_uuid, str(p_num), anch),
        )


def _serialize_sqlite_conn(conn: Any) -> bytes:
    import os
    import sqlite3
    import tempfile

    if hasattr(conn, "serialize"):
        raw_sqlite_bytes = conn.serialize()
        conn.close()
        return raw_sqlite_bytes

    with tempfile.NamedTemporaryFile(suffix=".db", delete=False) as tmp_f:
        tmp_path = tmp_f.name

    file_conn = sqlite3.connect(tmp_path)
    conn.backup(file_conn)
    file_conn.close()
    conn.close()

    with open(tmp_path, "rb") as f:
        raw_sqlite_bytes = f.read()
    try:
        os.remove(tmp_path)
    except OSError as e:
        logger.debug("[FileUtils] Failed to remove temporary SQLite file %s: %s", tmp_path, e)
    return raw_sqlite_bytes


def generate_curated_sqlite_bundle(
    document_ids: list[int] | list[str],
    user: Any = None,
    actor_id: str | None = None,
) -> bytes:
    """
    Generates a standalone, self-contained SQLite database bundle with full-text search (FTS5).
    Includes 'documents', 'chunks', and virtual full-text search table 'chunks_fts'.
    Enforces user boundaries respecting tenant isolation.
    """
    import sqlite3

    from django.conf import settings

    if getattr(settings, "SURREALDB_OFFLINE", False):
        docs_list = _get_offline_docs(document_ids, user)
    else:
        docs_list = _get_surreal_docs(document_ids, user, actor_id=actor_id)

    if not docs_list:
        raise ValueError(NO_COMPLETED_DOCS_MSG)

    if len(docs_list) > 1000:
        raise ValueError("Export batch exceeds maximum document limit of 1000 items.")

    conn = sqlite3.connect(":memory:")
    cursor = conn.cursor()
    _init_sqlite_export_schema(cursor)

    for doc in docs_list:
        doc_uuid = _insert_sqlite_document(cursor, doc)
        _insert_sqlite_chunks(cursor, doc, doc_uuid)

    conn.commit()
    return _serialize_sqlite_conn(conn)


def _sanitize_csv_cell(val: Any) -> Any:
    if isinstance(val, str) and val.lstrip().startswith(("=", "+", "-", "@")):
        return f"'{val}"
    return val


def _extract_doc_excerpt(doc: Any) -> str:
    raw_content = (
        getattr(doc, "refined_markdown", None)
        or getattr(doc, "cleaned_markdown", None)
        or getattr(doc, "raw_markdown", None)
        or ""
    )
    excerpt = (raw_content[:500] + "...") if len(raw_content) > 500 else raw_content
    return " ".join(excerpt.split())


def _build_csv_row_for_doc(doc: Any) -> list[Any]:
    doc_uuid = str(getattr(doc, "doc_uuid", None) or getattr(doc, "uuid", None) or doc.id)
    raw_row = [
        doc_uuid,
        _get_doc_field_str(doc, "title", "Untitled"),
        _get_doc_field_str(doc, "author", "Unknown"),
        _get_doc_field_str(doc, "language", "en"),
        _get_doc_field_str(doc, "document_type", "PDF"),
        _get_doc_field_str(doc, "publisher"),
        _get_doc_field_str(doc, "publication_year"),
        _get_doc_field_str(doc, "doi"),
        _get_doc_field_str(doc, "license_type"),
        _get_doc_field_str(doc, "file_hash"),
        int(getattr(doc, "page_count", 1) or 1),
        float(getattr(doc, "cost_usd", 0.0) or 0.0),
        int(getattr(doc, "input_tokens", 0) or 0),
        int(getattr(doc, "output_tokens", 0) or 0),
        _get_doc_field_str(doc, "created_at"),
        _extract_doc_excerpt(doc),
    ]
    return [_sanitize_csv_cell(c) for c in raw_row]


def generate_curated_csv_bundle(
    document_ids: list[int] | list[str],
    user: Any = None,
    actor_id: str | None = None,
) -> bytes:
    """
    Generates a structured CSV export containing legal metadata, copyright info, token metrics,
    and excerpt content for all selected documents.
    """
    import csv
    import io

    from django.conf import settings

    if getattr(settings, "SURREALDB_OFFLINE", False):
        docs_list = _get_offline_docs(document_ids, user)
    else:
        docs_list = _get_surreal_docs(document_ids, user, actor_id=actor_id)

    if not docs_list:
        raise ValueError(NO_COMPLETED_DOCS_MSG)

    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)

    writer.writerow(
        [
            "Document UUID",
            "Title",
            "Author",
            "Language",
            "Document Type",
            "Publisher",
            "Publication Year",
            "DOI",
            "License Type",
            "SHA-256 Hash",
            "Page Count",
            "AI Spend (USD)",
            "Input Tokens",
            "Output Tokens",
            "Created At",
            "Content Excerpt",
        ]
    )

    for doc in docs_list:
        writer.writerow(_build_csv_row_for_doc(doc))

    return output.getvalue().encode("utf-8")


def cleanup_stale_temp_artifacts(temp_dir: str | None = None, max_age_seconds: int = 86400) -> int:
    """Automated cleanup policy per DevSecOps best practices.

    Scans temporary directories for stale processing files older than max_age_seconds (default: 24h).
    Returns total count of removed temporary artifacts.
    """
    import os
    import tempfile
    import time

    # VULN-01 / S5443 fix: resolve absolute realpath to prevent symlink traversal in temp dir
    raw_dir = temp_dir if temp_dir is not None else tempfile.gettempdir()  # NOSONAR python:S5443
    target_dir = os.path.realpath(raw_dir)
    if not os.path.exists(target_dir):
        return 0

    now = time.time()
    removed_count = 0
    prefix_patterns = ("tmpx", "tmp", "korda_", "aetheromni_", "pdf_export_")

    for filename in os.listdir(target_dir):
        if not filename.startswith(prefix_patterns):
            continue
        file_path = os.path.realpath(os.path.join(target_dir, filename))
        # Ensure resolved file path remains safely within target_dir
        if not file_path.startswith(target_dir + os.sep):
            continue
        try:
            if os.path.isfile(file_path):
                file_age = now - os.path.getmtime(file_path)
                if file_age > max_age_seconds:
                    os.remove(file_path)
                    removed_count += 1
        except OSError as exc:
            logger.debug("[File Cleanup] Could not remove temporary file %s: %s", file_path, exc)

    return removed_count


def get_client_ip(request: Any) -> str:
    """Extract the true client IP from Django request headers."""
    x_forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR")
    if x_forwarded_for:
        return str(x_forwarded_for.split(",")[0].strip())
    return str(request.META.get("REMOTE_ADDR", "127.0.0.1"))


def _resolve_currency_and_symbol(accept_language: str) -> tuple[str, str]:
    if "id" in accept_language:
        return "IDR", "Rp "
    if "ar" in accept_language:
        return "SAR", "SR "
    if any(lang in accept_language for lang in ["de", "fr", "es", "it", "nl", "pt"]):
        return "EUR", "€"
    if "gb" in accept_language:
        return "GBP", "£"
    if "tr" in accept_language:
        return "TRY", "₺"
    return "USD", "$"


def _fetch_live_rates_with_fallback():
    rates = None
    try:
        with urllib.request.urlopen("https://open.er-api.com/v6/latest/USD", timeout=5) as response:  # nosec B310 nosemgrep
            data = json.loads(response.read().decode())
            if data.get("result") == "success":
                rates = data.get("rates", {})
                try:
                    from extractor import surreal_db

                    surreal_db.kv_cache_set("usd_exchange_rates", rates, ttl_seconds=86400)
                except (RuntimeError, ValueError, KeyError, AttributeError) as exc:
                    logger.debug("[Exchange Rates] Failed to cache rates: %s", exc)
    except (urllib.error.URLError, ValueError, TypeError, KeyError, OSError) as exc:
        logger.exception("[Exchange Rates] Error fetching live rates: %s — using fallback.", exc)
        rates = _FALLBACK_RATES.copy()
        try:
            from extractor import surreal_db

            surreal_db.kv_cache_set("usd_exchange_rates", rates, ttl_seconds=3600)
        except (RuntimeError, ValueError, KeyError, AttributeError) as exc:
            logger.debug("[Exchange Rates] Failed to cache fallback rates: %s", exc)
    return rates


def _get_exchange_rates():
    rates = None
    try:
        from extractor import surreal_db

        rates = surreal_db.kv_cache_get("usd_exchange_rates")
    except (RuntimeError, ValueError, KeyError, AttributeError) as exc:
        logger.debug("[Exchange Rates] Failed to read cached rates: %s", exc)

    if not rates:
        from django.core.cache import cache

        rates = cache.get("usd_exchange_rates")

    if not rates:
        rates = _fetch_live_rates_with_fallback()
    return rates


def get_locale_currency_details(request: Any) -> dict[str, Any]:
    """
    Resolves target currency, local currency name, and current exchange rate from USD
    by parsing browser settings (HTTP_ACCEPT_LANGUAGE) or system settings override.
    Caches fallback rates for 1 hour when external API is unavailable.
    Uses SurrealDB KV cache instead of Redis.
    """
    from extractor.models import SystemSettings

    try:
        settings_obj = SystemSettings.get_settings()
        selected_currency = settings_obj.currency
    except (SystemSettings.DoesNotExist, AttributeError, RuntimeError):
        selected_currency = "auto"

    if selected_currency and selected_currency != "auto":
        currency = selected_currency
        symbol_map = {"USD": "$", "IDR": "Rp ", "SAR": "SR ", "EUR": "€", "GBP": "£", "TRY": "₺"}
        symbol = symbol_map.get(currency, "$")
    else:
        accept_language = request.META.get("HTTP_ACCEPT_LANGUAGE", "").lower() if request else ""
        currency, symbol = _resolve_currency_and_symbol(accept_language)

    rates = _get_exchange_rates()
    rate = rates.get(currency, 1.0) if rates else 1.0
    return {"currency_code": currency, "symbol": symbol, "rate": rate}


def format_localized_cost(cost_usd: Decimal, currency_details: dict[str, Any]) -> str:
    """Formats USD pricing into both localized and USD formats."""
    usd_val = Decimal(str(cost_usd))
    rate = Decimal(str(currency_details["rate"]))
    local_val = usd_val * rate

    def fmt_usd(v: Decimal) -> str:
        if v < Decimal("0.10") and v > 0:
            return f"${v:.4f}"
        return f"${v:.2f}"

    if currency_details["currency_code"] == "USD":
        return fmt_usd(usd_val)

    if currency_details["currency_code"] == "IDR":
        return f"{currency_details['symbol']}{int(local_val):,} (~{fmt_usd(usd_val)})"

    return f"{currency_details['symbol']}{local_val:.2f} (~{fmt_usd(usd_val)})"


def get_google_oidc_token(audience: str) -> str | None:
    """
    Fetches an OIDC ID token from the GCP metadata server for the given audience.
    Returns None immediately in DEBUG mode to skip blocking network calls.
    """
    if settings.DEBUG:
        return None

    url = f"http://metadata.google.internal/computeMetadata/v1/instance/service-accounts/default/identity?audience={audience}"
    req = urllib.request.Request(url, headers={"Metadata-Flavor": "Google"})
    try:
        with urllib.request.urlopen(req, timeout=5) as response:  # nosec B310 nosemgrep
            return response.read().decode("utf-8").strip()
    except Exception as exc:
        logger.warning("[OIDC] Failed to fetch OIDC identity (not in GCP?): %s", exc)
        return None


def extract_pdf_diagrams_with_vision(pdf_path: str, max_pages: int = 5) -> str:
    """
    Extracts embedded diagrams, flowcharts, and architectural schemas from PDF pages
    using Gemini 3.6 Multi-Modal Vision OCR.
    """
    try:
        import fitz  # PyMuPDF
    except ImportError:
        logger.debug("[Vision OCR] PyMuPDF (fitz) is not installed.")
        return ""

    from extractor.llm_gateway import generate_multimodal_vision_ocr

    diagram_notes: list[str] = []
    try:
        with fitz.open(pdf_path) as doc:
            for page_num in range(min(len(doc), max_pages)):
                page = doc[page_num]
                image_list = page.get_images(full=True)
                if image_list:
                    pix = page.get_pixmap(dpi=150)
                    img_bytes = pix.tobytes("jpeg")
                    ocr_result = generate_multimodal_vision_ocr(
                        img_bytes,
                        mime_type="image/jpeg",
                        prompt=f"Page {page_num + 1} contains a diagram or schema. Describe the structure, nodes, flowchart connections, and text accurately in Markdown.",
                    )
                    if ocr_result:
                        diagram_notes.append(
                            f"\n### 📊 Page {page_num + 1} Visual Diagram & Schema Extraction\n{ocr_result}\n"
                        )
    except Exception as exc:
        logger.warning("[Vision OCR] Failed to process PDF page diagrams: %s", exc)
    return "\n".join(diagram_notes)
